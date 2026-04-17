#!/usr/bin/env python3
"""Build disbursement_daily_upload.sql from March xlsx + April CSV.

NO pandas. Uses openpyxl + csv stdlib.
Does NOT execute against the DB — writes SQL file only.
"""
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path("/Users/raghunandanmali/Downloads/DB/ESAF- Disb As on 31-03-2026.xlsx")
CSV_PATH = Path("/Users/raghunandanmali/Downloads/DB/Apirl.csv")
OUT_SQL = Path("/Users/raghunandanmali/Desktop/Coll_Db/data/disbursement_daily_upload.sql")

PRODUCT_MAP = {
    "204207": "IGL", "104207": "IGL", "264203": "IGL", "234008": "IGL",
    "604001": "FIG",
    "81402": "IL",
}

DDL = """-- disbursement_daily: date-level disbursement facts (ingested from March xlsx + April CSV)
CREATE TABLE IF NOT EXISTS disbursement_daily (
  id SERIAL PRIMARY KEY,
  disb_date DATE NOT NULL,
  region_name VARCHAR,
  district_name VARCHAR,
  branch_name VARCHAR,
  emp_id VARCHAR,
  officer_name VARCHAR,
  product_name VARCHAR NOT NULL,
  disb_count INTEGER NOT NULL,
  disb_amount NUMERIC NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_disb_daily_date ON disbursement_daily(disb_date);
CREATE INDEX IF NOT EXISTS idx_disb_daily_branch ON disbursement_daily(branch_name);
CREATE INDEX IF NOT EXISTS idx_disb_daily_emp ON disbursement_daily(emp_id);
"""


def sql_str(v):
    if v is None or v == "":
        return "NULL"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def sql_date(d):
    if isinstance(d, (datetime, date)):
        return f"'{d.strftime('%Y-%m-%d')}'"
    return "NULL"


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clean_officer(name):
    """Strip trailing '(code' or '(code)' from CreditOffcierName."""
    if not name:
        return None
    return re.sub(r"\s*\([^)]*\)?\s*$", "", str(name)).strip() or None


def extract_code(name):
    """Extract numeric code inside (NNN...) from CreditOffcierName."""
    if not name:
        return None
    m = re.search(r"\((\d+)", str(name))
    return m.group(1) if m else None


def map_product(product_id, product_name=None):
    if product_name and product_name in ("IGL", "FIG", "IL"):
        return product_name
    pid = str(product_id).strip().split(".")[0] if product_id else ""
    return PRODUCT_MAP.get(pid)


def read_xlsx():
    """Return (aggregates, branch_lookup, code_to_emp_lookup, stats)."""
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Data"]
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    idx = {h: i for i, h in enumerate(headers) if h}

    agg = defaultdict(lambda: [0, 0.0])  # key -> [count, sum_amt]
    branch_lookup = {}  # UPPER(branch_name) -> (region, division, area)
    code_to_emp = {}    # numeric code suffix -> emp_id
    skipped = defaultdict(int)
    total = 0
    dates = []

    for row in it:
        total += 1
        g = lambda k: row[idx[k]] if k in idx and idx[k] < len(row) else None
        d = parse_date(g("LoanDisbDate"))
        if not d:
            skipped["bad_date"] += 1
            continue
        product = map_product(g("SchemeID/ProductID"), g("Product Name"))
        if not product:
            skipped["bad_product"] += 1
            continue
        amount = g("DisbAmount") or 0
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            skipped["bad_amount"] += 1
            continue

        region = (g("New Region") or "").strip() or None
        division = (g("New Division") or "").strip() or None
        area = (g("New Area") or "").strip() or None
        branch = (g("BranchName") or "").strip() or None
        emp_id = (g("EMP ID") or "").strip() or None
        officer_raw = g("CreditOffcierName")
        officer = clean_officer(officer_raw)
        code = extract_code(officer_raw)

        if branch:
            key = branch.upper()
            if key not in branch_lookup:
                branch_lookup[key] = (region, division, area)
        if code and emp_id:
            code_to_emp[code] = emp_id

        key = (d, region, division, branch, emp_id, officer, product)
        agg[key][0] += 1
        agg[key][1] += amount
        dates.append(d)

    stats = {"total": total, "skipped": dict(skipped), "dates": dates}
    return agg, branch_lookup, code_to_emp, stats


def read_csv(branch_lookup, code_to_emp):
    agg = defaultdict(lambda: [0, 0.0])
    skipped = defaultdict(int)
    total = 0
    dates = []
    unmatched_branches = set()
    unmatched_codes = 0

    with CSV_PATH.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total += 1
            d = parse_date(row.get("LoanDisbDate"))
            if not d:
                skipped["bad_date"] += 1
                continue
            product = map_product(row.get("SchemeID/ProductID"))
            if not product:
                skipped["bad_product"] += 1
                continue
            amount_raw = row.get("DisbAmount") or "0"
            try:
                amount = float(str(amount_raw).replace(",", ""))
            except ValueError:
                skipped["bad_amount"] += 1
                continue

            branch = (row.get("BranchName") or "").strip() or None
            officer_raw = row.get("CreditOffcierName")
            officer = clean_officer(officer_raw)
            code = extract_code(officer_raw)

            region = division = area = None
            if branch:
                hit = branch_lookup.get(branch.upper())
                if hit:
                    region, division, area = hit
                else:
                    unmatched_branches.add(branch)

            emp_id = None
            if code:
                emp_id = code_to_emp.get(code)
                if not emp_id:
                    unmatched_codes += 1

            key = (d, region, division, branch, emp_id, officer, product)
            agg[key][0] += 1
            agg[key][1] += amount
            dates.append(d)

    stats = {
        "total": total,
        "skipped": dict(skipped),
        "dates": dates,
        "unmatched_branches": len(unmatched_branches),
        "unmatched_codes": unmatched_codes,
    }
    return agg, stats


def write_sql(xlsx_agg, csv_agg):
    lines = [DDL, ""]
    lines.append("BEGIN;")
    lines.append("")

    cols = "(disb_date, region_name, district_name, branch_name, emp_id, officer_name, product_name, disb_count, disb_amount)"

    def emit_block(title, agg):
        lines.append(f"-- {title}: {len(agg)} aggregated rows")
        if not agg:
            return
        batch = []
        for (d, region, division, branch, emp_id, officer, product), (cnt, amt) in agg.items():
            vals = (
                f"({sql_date(d)}, {sql_str(region)}, {sql_str(division)}, "
                f"{sql_str(branch)}, {sql_str(emp_id)}, {sql_str(officer)}, "
                f"{sql_str(product)}, {cnt}, {amt:.2f})"
            )
            batch.append(vals)
            if len(batch) >= 500:
                lines.append(f"INSERT INTO disbursement_daily {cols} VALUES")
                lines.append(",\n".join(batch) + ";")
                batch = []
        if batch:
            lines.append(f"INSERT INTO disbursement_daily {cols} VALUES")
            lines.append(",\n".join(batch) + ";")
        lines.append("")

    emit_block("March xlsx rows", xlsx_agg)
    emit_block("April CSV rows", csv_agg)
    lines.append("COMMIT;")
    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")


def main():
    print("Reading xlsx ...", file=sys.stderr)
    xlsx_agg, branch_lookup, code_to_emp, xstats = read_xlsx()
    print(f"  xlsx raw rows: {xstats['total']}, aggregated: {len(xlsx_agg)}, skipped: {xstats['skipped']}")
    if xstats["dates"]:
        print(f"  xlsx date range: {min(xstats['dates'])} → {max(xstats['dates'])}")
    print(f"  branch_lookup entries: {len(branch_lookup)}, code→emp entries: {len(code_to_emp)}")

    print("Reading csv ...", file=sys.stderr)
    csv_agg, cstats = read_csv(branch_lookup, code_to_emp)
    print(f"  csv raw rows: {cstats['total']}, aggregated: {len(csv_agg)}, skipped: {cstats['skipped']}")
    if cstats["dates"]:
        print(f"  csv date range: {min(cstats['dates'])} → {max(cstats['dates'])}")
    print(f"  csv unmatched_branches: {cstats['unmatched_branches']}, unmatched_codes: {cstats['unmatched_codes']}")

    write_sql(xlsx_agg, csv_agg)
    print(f"Wrote SQL → {OUT_SQL}")


if __name__ == "__main__":
    main()

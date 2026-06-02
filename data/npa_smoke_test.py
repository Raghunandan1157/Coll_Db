#!/usr/bin/env python3
"""NPA upload smoke test — no HTTP, no deploy, no commit.

Mirrors the /api/npa/upload pipeline in server/index.js (NPA_HEADER_MAP,
parsers, bulk insert into npa_activation_runs + npa_activation_rows) then
ROLLS BACK the transaction. Uses the project's direct psycopg2 pattern
(52.66.163.52, same creds as migrate_v2.py).

Reports:
  - rows would-be-inserted (N)
  - distinct branches (D) and mapped-to-v2 (M/D)
  - up to 5 sample unmapped branch names
  - DDL alignment issues (VARCHAR overflow, bad types)
  - by-region rollup from the in-TX rows
"""
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values, Json

XLSX_PATH = Path(
    "/Users/raghunandanmali/Desktop/OFFICE WORK/NPA_Activation Report/v2/"
    "v3 - Database/processed/NPA collection Task__merged_20260423_123124.xlsx"
)

DB = {
    "host": "52.66.163.52",
    "port": 5432,
    "user": "Raghunandan1157",
    "password": "raghu",
    "dbname": "postgres",
}

EXECUTE = True   # True = open TX, insert, rollup, ROLLBACK. False = dry-run (no DB writes).
BATCH = 200

# Mirrors NPA_HEADER_MAP in server/index.js L834-880 (keys normalized: [a-z0-9]+).
NPA_HEADER_MAP = {
    "identifier": "identifier",
    "designation": "designation",
    "memberid": "member_id",
    "accountid": "account_id",
    "mobile": "mobile",
    "groupname": "group_name",
    "productid": "product_id",
    "loanamount": "loan_amount",
    "loanoustanding": "loan_outstanding",
    "odamount": "od_amount",
    "dpddays": "dpd_bucket",
    "type": "task_type",
    "employee": "employee",
    "title": "title",
    "started": "started_at",
    "completed": "completed_on",
    "timetaken": "time_taken",
    "priority": "priority",
    "status": "status",
    "presentstate": "present_state",
    "customer": "customer",
    "customeraddress": "customer_address",
    "taskaddress": "task_address",
    "checkin": "check_in_address",
    "checkout": "check_out_address",
    "taskcompletelatlng": "task_latlng",
    "followup": "follow_up",
    "nextfollowuptime": "next_follow_up_at",
    "followupcomment": "follow_up_comment",
    "customerloanid": "customer_loan_id",
    "memberalternativecontactnumber": "alt_contact",
    "memberattendence": "member_attendance",
    "memberphoto": "member_photo",
    "paymentstatus": "payment_status",
    "promisetopaydateandtime": "promise_to_pay_at",
    "paymentmode": "payment_mode",
    "amount": "amount",
    "receiptnumber": "receipt_number",
    "utrnumber": "utr_number",
    "reasonfornotcollected": "reason_not_collected",
    "ifmemberabsenthousephoto": "absent_house_photo",
    "transactionscreenshot": "transaction_screenshot",
    "loanstatus": "loan_status",
    "collectionreport": "collection_report",
    "worklocation": "branch_name",
}

BOOL_COLS = {"follow_up", "member_photo", "absent_house_photo", "transaction_screenshot"}
NUMERIC_COLS = {"loan_amount", "loan_outstanding", "od_amount", "amount", "collection_report"}
BIGINT_COLS = {"account_id", "customer_loan_id"}
DATE_COLS = {"completed_on"}
TS_COLS = {"started_at", "next_follow_up_at", "promise_to_pay_at"}

# Destination column order (matches dbCols in server/index.js L967-973).
DB_COLS = [
    "run_id", "branch_name", "identifier", "designation", "member_id", "account_id",
    "mobile", "group_name", "product_id", "loan_amount", "loan_outstanding", "od_amount",
    "dpd_bucket", "task_type", "employee", "title", "started_at", "completed_on",
    "time_taken", "priority", "status", "present_state", "customer", "customer_address",
    "task_address", "check_in_address", "check_out_address", "task_latlng", "follow_up",
    "next_follow_up_at", "follow_up_comment", "customer_loan_id", "alt_contact",
    "member_attendance", "member_photo", "payment_status", "promise_to_pay_at",
    "payment_mode", "amount", "receipt_number", "utr_number", "reason_not_collected",
    "absent_house_photo", "transaction_screenshot", "loan_status", "collection_report",
]


def _key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def parse_bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    if s in ("yes", "true", "1", "y"):
        return True
    if s in ("no", "false", "0", "n"):
        return False
    return None


def parse_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.upper() == "NA" or s == "-":
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def parse_bigint(v):
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s or s.upper() == "NA":
        return None
    cleaned = s.replace(",", "").split(".")[0]
    if not re.fullmatch(r"-?\d+", cleaned):
        return None
    return int(cleaned)


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[: len(fmt.replace("%Y", "2026").replace("%m", "01").replace("%d", "01").replace("%H", "00").replace("%M", "00").replace("%S", "00"))], fmt).date()
        except ValueError:
            continue
    # Try JS-style Date() best-effort parse for ISO-ish strings
    try:
        return datetime.fromisoformat(s[:19].replace("T", " ")).date()
    except ValueError:
        pass
    return None


def parse_ts(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Format from merged xlsx: "2026-04-23 10:20:43 +0530"
    m = re.match(r"^(\d{4}-\d{2}-\d{2})[ T](\d{2}:\d{2}:\d{2})(?:\s*([+-]\d{4}))?", s)
    if m:
        ymd, hms, tz = m.groups()
        iso = f"{ymd}T{hms}"
        if tz:
            iso += f"{tz[:3]}:{tz[3:]}"
        try:
            return datetime.fromisoformat(iso)
        except ValueError:
            pass
    # Short form "2026-04-28 10:19" for promise_to_pay_at
    try:
        return datetime.fromisoformat(s.replace("T", " ")[:19])
    except ValueError:
        return None


def shape_row(row, col_map):
    rec = {}
    for c, raw in enumerate(row):
        if c >= len(col_map):
            break
        target = col_map[c]
        if not target:
            continue
        if target == "branch_name":
            val = None if raw is None else str(raw).strip().upper() or None
        elif target in BOOL_COLS:
            val = parse_bool(raw)
        elif target in BIGINT_COLS:
            val = parse_bigint(raw)
        elif target in NUMERIC_COLS:
            val = parse_num(raw)
        elif target in DATE_COLS:
            val = parse_date(raw)
        elif target in TS_COLS:
            val = parse_ts(raw)
        else:
            val = None if (raw is None or str(raw).strip() == "") else str(raw)
        rec[target] = val
    return rec


# VARCHAR column limits from npa_activation_rows DDL.
VARCHAR_LIMITS = {
    "branch_name": 100, "identifier": 20, "designation": 100, "member_id": 30,
    "mobile": 20, "group_name": 150, "product_id": 30, "dpd_bucket": 40,
    "task_type": 50, "employee": 150, "title": 60, "priority": 20, "status": 30,
    "alt_contact": 50, "member_attendance": 20, "payment_status": 30,
    "payment_mode": 40, "receipt_number": 50, "utr_number": 50, "loan_status": 30,
}


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"[load] {XLSX_PATH.name}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    # Backend picks sheet matching /NPA Collection Task/i, else first.
    sheet_name = next(
        (s for s in wb.sheetnames if re.search(r"npa collection task", s, re.I)),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    print(f"[load] sheet={sheet_name!r}")

    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    col_map = [NPA_HEADER_MAP.get(_key(h)) for h in header]
    mapped_hdrs = [(h, col_map[i]) for i, h in enumerate(header) if col_map[i]]
    unmapped_hdrs = [h for i, h in enumerate(header) if not col_map[i]]
    print(f"[hdr] mapped={len(mapped_hdrs)}, unmapped={len(unmapped_hdrs)}")
    if unmapped_hdrs:
        print(f"[hdr] unmapped columns (dropped by backend): {unmapped_hdrs}")

    parsed = []
    skipped_branch = 0
    skipped_acct = 0
    npa_count = 0
    max_completed = None
    branches_seen = set()

    for row in rows_iter:
        if not row or all(v is None for v in row):
            continue
        rec = shape_row(row, col_map)
        if not rec.get("branch_name"):
            skipped_branch += 1
            continue
        if rec.get("account_id") is None:
            skipped_acct += 1
            continue
        branches_seen.add(rec["branch_name"])
        if str(rec.get("loan_status") or "").upper() == "NPA":
            npa_count += 1
        co = rec.get("completed_on")
        if co and (max_completed is None or co > max_completed):
            max_completed = co
        parsed.append(rec)
    wb.close()

    print(f"[parse] rows_kept={len(parsed)} skipped_no_branch={skipped_branch} skipped_no_account={skipped_acct}")
    print(f"[parse] npa={npa_count} distinct_branches={len(branches_seen)} max_completed={max_completed}")

    # DDL alignment pre-check: VARCHAR overflow + NUMERIC(14,2) overflow (|v| < 1e12).
    overflow = defaultdict(list)
    for rec in parsed:
        for col, limit in VARCHAR_LIMITS.items():
            v = rec.get(col)
            if v is not None and isinstance(v, str) and len(v) > limit:
                if len(overflow[col]) < 3:
                    overflow[col].append((len(v), v[:60]))
    numeric_overflow = defaultdict(list)
    for rec in parsed:
        for col in NUMERIC_COLS:
            v = rec.get(col)
            if v is not None and abs(v) >= 1e12:
                if len(numeric_overflow[col]) < 3:
                    numeric_overflow[col].append((v, rec.get("account_id")))
                rec[col] = None  # sanitize so INSERT survives
    if overflow:
        print("[ddl] VARCHAR OVERFLOW DETECTED:")
        for col, samples in overflow.items():
            print(f"  - {col} (limit={VARCHAR_LIMITS[col]}): {samples}")
    else:
        print("[ddl] no VARCHAR overflow")
    if numeric_overflow:
        print("[ddl] NUMERIC(14,2) OVERFLOW (|v| >= 1e12) — sanitized to NULL for insert:")
        for col, samples in numeric_overflow.items():
            cnt = sum(1 for r in parsed if False)  # placeholder; count below
            total = sum(
                1
                for r in parsed
                # counted again by rescanning original, but we already nulled; keep sample list truthful
            )
            print(f"  - {col} (limit numeric(14,2)): samples={samples}")
    else:
        print("[ddl] no NUMERIC(14,2) overflow")

    if not EXECUTE:
        print("[dry-run] EXECUTE=False — stopping before DB connection")
        return

    print(f"[db] connecting to {DB['host']}:{DB['port']}/{DB['dbname']} as {DB['user']}")
    conn = psycopg2.connect(**DB, connect_timeout=15)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            # Pre-insert branch match rate (v2_branches is reference-only, safe outside TX).
            cur.execute("SELECT UPPER(branch_name) FROM v2_branches")
            v2_set = {r[0] for r in cur.fetchall()}
            mapped = branches_seen & v2_set
            unmapped = branches_seen - v2_set
            print(f"[v2] v2_branches={len(v2_set)} mapped={len(mapped)}/{len(branches_seen)} unmapped={len(unmapped)}")
            if unmapped:
                print(f"[v2] sample unmapped: {sorted(unmapped)[:5]}")

            summary = {
                "final_rows": len(parsed),
                "branches_seen": len(branches_seen),
                "npa_seen": npa_count,
                "sheet": sheet_name,
                "smoke_test": True,
            }
            cur.execute(
                """INSERT INTO npa_activation_runs
                   (source_filename, report_date, row_count, npa_count, summary)
                   VALUES (%s, %s, %s, %s, %s) RETURNING run_id""",
                (XLSX_PATH.name, max_completed, len(parsed), npa_count, Json(summary)),
            )
            run_id = cur.fetchone()[0]
            print(f"[tx] run_id={run_id} (will be rolled back)")

            col_list = ", ".join(DB_COLS)
            placeholders = "(" + ", ".join(["%s"] * len(DB_COLS)) + ")"

            def tuplify(rec):
                return tuple(run_id if c == "run_id" else rec.get(c) for c in DB_COLS)

            inserted = 0
            for i in range(0, len(parsed), BATCH):
                chunk = parsed[i : i + BATCH]
                vals = [tuplify(r) for r in chunk]
                execute_values(
                    cur,
                    f"INSERT INTO npa_activation_rows ({col_list}) VALUES %s",
                    vals,
                    template=placeholders,
                    page_size=BATCH,
                )
                inserted += len(chunk)
                if inserted % 2000 == 0 or inserted == len(parsed):
                    print(f"[tx] inserted {inserted}/{len(parsed)}")

            # Verify rowcount within TX.
            cur.execute(
                "SELECT COUNT(*) FROM npa_activation_rows WHERE run_id = %s", (run_id,)
            )
            confirmed = cur.fetchone()[0]
            print(f"[tx] npa_activation_rows count(run_id={run_id}): {confirmed}")

            # By-region rollup against the in-TX data (same SQL as /api/npa/by-region).
            cur.execute(
                """
                SELECT COALESCE(s.state_name, '(unmapped)') AS region,
                       COUNT(*) FILTER (WHERE nar.loan_status = 'NPA')::int AS npa_count,
                       COALESCE(SUM(nar.collection_report), 0)::numeric AS collection_sum,
                       COUNT(DISTINCT nar.branch_name)::int AS branch_count,
                       COUNT(DISTINCT nar.account_id)::int AS account_count,
                       COUNT(*)::int AS total_rows
                  FROM npa_activation_rows nar
             LEFT JOIN v2_branches  b  ON UPPER(b.branch_name) = UPPER(nar.branch_name)
             LEFT JOIN v2_areas     a  ON b.area_id = a.area_id
             LEFT JOIN v2_divisions dv ON a.division_id = dv.division_id
             LEFT JOIN v2_states    s  ON dv.state_id = s.state_id
                 WHERE nar.run_id = %s
              GROUP BY COALESCE(s.state_name, '(unmapped)')
              ORDER BY region
                """,
                (run_id,),
            )
            print("[rollup] by-region:")
            print(f"  {'region':<30s} {'rows':>8s} {'npa':>8s} {'branches':>10s} {'accounts':>10s} {'coll_sum':>14s}")
            total_rows = total_npa = 0
            unmapped_rows = 0
            for region, nc, cs, bc, ac, tr in cur.fetchall():
                total_rows += tr
                total_npa += nc
                if region == "(unmapped)":
                    unmapped_rows = tr
                print(f"  {region:<30s} {tr:>8d} {nc:>8d} {bc:>10d} {ac:>10d} {cs:>14.2f}")
            print(f"[rollup] totals: rows={total_rows} npa={total_npa} unmapped_rows={unmapped_rows}")
            if total_rows:
                print(f"[rollup] branch join rate: {(total_rows - unmapped_rows)}/{total_rows} = {100*(total_rows-unmapped_rows)/total_rows:.2f}%")

        # NEVER commit — smoke test only.
        conn.rollback()
        print("[tx] ROLLED BACK — DB is clean")
    except Exception as e:
        conn.rollback()
        print(f"[tx] ERROR (rolled back): {type(e).__name__}: {e}", file=sys.stderr)
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
